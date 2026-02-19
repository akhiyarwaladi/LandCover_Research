"""
Generate Statistical Analysis Tables and Figures

Produces statistical tests and temporal analysis outputs:
- McNemar's test (pairwise approach comparison)
- Cohen's Kappa analysis
- Annual deforestation statistics
- Temporal trend analysis

Usage:
    python scripts/generate_statistical_analysis.py

Output:
    results/tables/statistical/*.xlsx
    results/tables/change_analysis/*.xlsx
    results/figures/statistical/*.png
    results/figures/temporal_analysis/*.png
"""

import os
import sys
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.visualizer import (
    plot_annual_deforestation_trend,
    plot_cumulative_loss,
)


# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_DIR = os.path.join(BASE_DIR, 'results')
STAT_TABLES_DIR = os.path.join(RESULTS_DIR, 'tables', 'statistical')
CHANGE_TABLES_DIR = os.path.join(RESULTS_DIR, 'tables', 'change_analysis')
STAT_FIGURES_DIR = os.path.join(RESULTS_DIR, 'figures', 'statistical')
TEMPORAL_FIGURES_DIR = os.path.join(RESULTS_DIR, 'figures', 'temporal_analysis')

APPROACHES = {
    'PCC-ResNet101': 'pcc_resnet101',
    'Siamese-ResNet50': 'siamese_resnet50',
    'RF-Change': 'rf_change',
}


def format_excel(writer, sheet_name):
    """Apply professional formatting."""
    worksheet = writer.sheets[sheet_name]
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    for column in worksheet.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(max_length + 4, 12)


def load_results():
    """Load all approach results."""
    results = {}
    for name, model_dir in APPROACHES.items():
        path = os.path.join(RESULTS_DIR, 'models', model_dir, 'test_results.npz')
        if os.path.exists(path):
            data = np.load(path, allow_pickle=True)
            results[name] = {key: data[key] for key in data.files}
    return results


def mcnemar_test(targets, preds_a, preds_b):
    """
    Perform McNemar's test between two classifiers.

    Returns: (chi_squared, p_value)
    """
    correct_a = (preds_a == targets)
    correct_b = (preds_b == targets)

    # Contingency table
    b = np.sum(correct_a & ~correct_b)  # A correct, B wrong
    c = np.sum(~correct_a & correct_b)  # A wrong, B correct

    if b + c == 0:
        return 0.0, 1.0

    chi2 = (abs(b - c) - 1) ** 2 / (b + c)
    p_value = 1 - scipy_stats.chi2.cdf(chi2, df=1)

    return float(chi2), float(p_value)


def generate_mcnemar_table(all_results):
    """Generate pairwise McNemar's test table."""
    print("  Generating McNemar's test table...")

    names = list(all_results.keys())
    if len(names) < 2:
        print("    Need at least 2 approaches for comparison")
        return

    # Use the first available targets
    targets = None
    for r in all_results.values():
        targets = r['targets']
        break

    rows = []
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            preds_a = all_results[names[i]]['predictions']
            preds_b = all_results[names[j]]['predictions']

            # Ensure same length
            min_len = min(len(targets), len(preds_a), len(preds_b))
            chi2, p_val = mcnemar_test(
                targets[:min_len], preds_a[:min_len], preds_b[:min_len]
            )

            sig = '***' if p_val < 0.001 else '**' if p_val < 0.01 else \
                  '*' if p_val < 0.05 else 'ns'

            rows.append({
                'Comparison': f'{names[i]} vs {names[j]}',
                'Chi-squared': round(chi2, 4),
                'p-value': round(p_val, 6),
                'Significance': sig,
            })

    df = pd.DataFrame(rows)

    output_path = os.path.join(STAT_TABLES_DIR, 'mcnemar_pairwise.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='McNemar', index=False)
        format_excel(writer, 'McNemar')

    print(f"    Saved: {output_path}")
    return df


def generate_kappa_table(all_results):
    """Generate Kappa coefficient analysis table."""
    print("  Generating Kappa analysis table...")

    rows = []
    for name, results in all_results.items():
        kappa = float(results.get('kappa', 0))

        if kappa < 0:
            interpretation = 'Poor'
        elif kappa < 0.2:
            interpretation = 'Slight'
        elif kappa < 0.4:
            interpretation = 'Fair'
        elif kappa < 0.6:
            interpretation = 'Moderate'
        elif kappa < 0.8:
            interpretation = 'Substantial'
        else:
            interpretation = 'Almost Perfect'

        rows.append({
            'Approach': name,
            'Kappa': round(kappa, 4),
            'Interpretation': interpretation,
            'Accuracy': round(float(results.get('accuracy', 0)), 4),
            'F1-Macro': round(float(results.get('f1_macro', 0)), 4),
        })

    df = pd.DataFrame(rows)

    output_path = os.path.join(STAT_TABLES_DIR, 'kappa_analysis.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Kappa', index=False)
        format_excel(writer, 'Kappa')

    print(f"    Saved: {output_path}")


def generate_annual_deforestation_table():
    """Generate annual deforestation statistics table from change labels."""
    print("  Generating annual deforestation statistics...")

    import rasterio

    labels_dir = os.path.join(BASE_DIR, 'data', 'change_labels', 'annual')
    if not os.path.exists(labels_dir):
        print("    Change labels not found. Run prepare_change_labels.py first.")
        return None

    rows = []
    cumulative = 0

    for year in range(2018, 2025):
        label_path = os.path.join(labels_dir, f'change_{year}.tif')
        if not os.path.exists(label_path):
            continue

        with rasterio.open(label_path) as src:
            data = src.read(1)

        n_pixels = int(np.sum(data > 0))
        area_ha = n_pixels * 0.04
        area_km2 = area_ha / 100
        total_pixels = data.size
        rate_pct = (n_pixels / total_pixels) * 100
        cumulative += area_ha

        rows.append({
            'Year': year,
            'Deforestation (pixels)': n_pixels,
            'Area (ha)': round(area_ha, 1),
            'Area (km2)': round(area_km2, 2),
            'Rate (%)': round(rate_pct, 4),
            'Cumulative (ha)': round(cumulative, 1),
        })

    if not rows:
        return None

    df = pd.DataFrame(rows)

    output_path = os.path.join(CHANGE_TABLES_DIR, 'annual_deforestation.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Annual', index=False)
        format_excel(writer, 'Annual')

    print(f"    Saved: {output_path}")

    # Create stats dict for visualization
    stats = {
        'years': [r['Year'] for r in rows],
        'area_ha': [r['Area (ha)'] for r in rows],
        'area_km2': [r['Area (km2)'] for r in rows],
        'rate_pct': [r['Rate (%)'] for r in rows],
        'cumulative_ha': [r['Cumulative (ha)'] for r in rows],
    }

    # Add trend
    if len(stats['years']) >= 3:
        x = np.array(stats['years'], dtype=float)
        y = np.array(stats['area_ha'], dtype=float)
        slope, intercept, r_value, p_value, _ = scipy_stats.linregress(x, y)
        stats['trend'] = {
            'slope_ha_per_year': slope,
            'intercept': intercept,
            'r_squared': r_value ** 2,
            'p_value': p_value,
        }
    else:
        stats['trend'] = None

    return stats


def generate_temporal_figures(stats):
    """Generate temporal analysis figures."""
    if stats is None:
        return

    print("  Generating temporal analysis figures...")

    # Annual deforestation trend
    save_path = os.path.join(TEMPORAL_FIGURES_DIR, 'annual_deforestation_trend.png')
    plot_annual_deforestation_trend(stats, save_path=save_path)

    # Cumulative loss
    save_path = os.path.join(TEMPORAL_FIGURES_DIR, 'cumulative_forest_loss.png')
    plot_cumulative_loss(stats, save_path=save_path)


def generate_mcnemar_heatmap(all_results):
    """Generate McNemar's test p-value heatmap."""
    import matplotlib.pyplot as plt
    import seaborn as sns

    names = list(all_results.keys())
    if len(names) < 2:
        return

    targets = list(all_results.values())[0]['targets']
    n = len(names)
    p_matrix = np.ones((n, n))

    for i in range(n):
        for j in range(i + 1, n):
            preds_a = all_results[names[i]]['predictions']
            preds_b = all_results[names[j]]['predictions']
            min_len = min(len(targets), len(preds_a), len(preds_b))
            _, p_val = mcnemar_test(targets[:min_len], preds_a[:min_len], preds_b[:min_len])
            p_matrix[i, j] = p_val
            p_matrix[j, i] = p_val

    fig, ax = plt.subplots(figsize=(8, 6), facecolor='white')

    mask = np.zeros_like(p_matrix, dtype=bool)
    np.fill_diagonal(mask, True)

    sns.heatmap(p_matrix, annot=True, fmt='.4f', cmap='RdYlGn',
                xticklabels=names, yticklabels=names,
                mask=mask, ax=ax, vmin=0, vmax=0.1,
                cbar_kws={'label': 'p-value'})

    ax.set_title("McNemar's Test p-values", fontsize=14, fontweight='bold')
    plt.tight_layout()

    save_path = os.path.join(STAT_FIGURES_DIR, 'mcnemar_pvalue_matrix.png')
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"    Saved: {save_path}")


def main():
    """Generate all statistical analysis outputs."""
    print("=" * 60)
    print("GENERATE STATISTICAL ANALYSIS")
    print("=" * 60)

    for d in [STAT_TABLES_DIR, CHANGE_TABLES_DIR, STAT_FIGURES_DIR, TEMPORAL_FIGURES_DIR]:
        os.makedirs(d, exist_ok=True)

    # Load results
    print("\nLoading approach results...")
    all_results = load_results()
    print(f"  Loaded {len(all_results)} approaches")

    # Statistical tests
    if len(all_results) >= 2:
        print("\n--- McNemar's Test ---")
        generate_mcnemar_table(all_results)
        generate_mcnemar_heatmap(all_results)

    if all_results:
        print("\n--- Kappa Analysis ---")
        generate_kappa_table(all_results)

    # Temporal analysis
    print("\n--- Annual Deforestation Analysis ---")
    stats = generate_annual_deforestation_table()

    print("\n--- Temporal Figures ---")
    generate_temporal_figures(stats)

    print("\n" + "=" * 60)
    print("STATISTICAL ANALYSIS COMPLETE")
    print("=" * 60)
    print(f"  Statistical tables: {STAT_TABLES_DIR}")
    print(f"  Change analysis: {CHANGE_TABLES_DIR}")
    print(f"  Statistical figures: {STAT_FIGURES_DIR}")
    print(f"  Temporal figures: {TEMPORAL_FIGURES_DIR}")


if __name__ == '__main__':
    main()
