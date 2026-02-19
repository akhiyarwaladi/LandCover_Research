"""
Generate Publication-Ready Comparison Tables and Figures

Compares all 3 change detection approaches and generates:
- Excel tables with performance metrics
- Confusion matrix figures
- Training curve figures
- Approach comparison bar charts

Usage:
    python scripts/generate_publication_comparison.py

Input:
    results/models/pcc_resnet101/test_results.npz
    results/models/siamese_resnet50/test_results.npz
    results/models/rf_change/test_results.npz

Output:
    results/tables/performance/*.xlsx
    results/figures/confusion_matrices/*.png
    results/figures/training_curves/*.png
"""

import os
import sys
import json
import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.visualizer import (
    plot_confusion_matrices_comparison,
    plot_training_curves,
    plot_approach_comparison_bar,
    plot_feature_importance,
)


# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_DIR = os.path.join(BASE_DIR, 'results')
TABLES_DIR = os.path.join(RESULTS_DIR, 'tables', 'performance')
FIGURES_DIR = os.path.join(RESULTS_DIR, 'figures')

APPROACHES = {
    'PCC-ResNet101': 'pcc_resnet101',
    'Siamese-ResNet50': 'siamese_resnet50',
    'RF-Change': 'rf_change',
}


def load_approach_results(approach_name, model_dir):
    """Load test results for an approach."""
    results_path = os.path.join(RESULTS_DIR, 'models', model_dir, 'test_results.npz')

    if not os.path.exists(results_path):
        print(f"  WARNING: Results not found for {approach_name}")
        return None

    data = np.load(results_path, allow_pickle=True)
    return {key: data[key] for key in data.files}


def format_excel_table(writer, sheet_name):
    """Apply professional formatting to an Excel sheet."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max(max_length + 4, 12)


def generate_performance_table(all_results):
    """Generate overall performance comparison table."""
    rows = []
    for name, results in all_results.items():
        if results is None:
            continue
        rows.append({
            'Approach': name,
            'Accuracy': float(results.get('accuracy', 0)),
            'F1-Macro': float(results.get('f1_macro', 0)),
            'F1-Weighted': float(results.get('f1_weighted', 0)),
            'F1-Change': float(results.get('f1_change', 0)),
            'Kappa': float(results.get('kappa', 0)),
        })

    if not rows:
        print("  No results to generate table")
        return

    df = pd.DataFrame(rows)

    output_path = os.path.join(TABLES_DIR, 'approach_comparison.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Comparison', index=False)
        format_excel_table(writer, 'Comparison')

    print(f"  Saved: {output_path}")
    return df


def generate_per_class_table(all_results):
    """Generate per-class performance table."""
    from sklearn.metrics import precision_recall_fscore_support

    rows = []
    class_names = ['No Change', 'Deforestation']

    for name, results in all_results.items():
        if results is None:
            continue

        targets = results['targets']
        predictions = results['predictions']

        precision, recall, f1, support = precision_recall_fscore_support(
            targets, predictions, labels=[0, 1], zero_division=0
        )

        for i, cls_name in enumerate(class_names):
            rows.append({
                'Approach': name,
                'Class': cls_name,
                'Precision': float(precision[i]),
                'Recall': float(recall[i]),
                'F1-Score': float(f1[i]),
                'Support': int(support[i]),
            })

    if not rows:
        return

    df = pd.DataFrame(rows)

    output_path = os.path.join(TABLES_DIR, 'per_class_performance.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Per-Class', index=False)
        format_excel_table(writer, 'Per-Class')

    print(f"  Saved: {output_path}")


def generate_confusion_matrices(all_results):
    """Generate confusion matrix comparison figure."""
    results_with_cm = {}
    for name, results in all_results.items():
        if results is not None and 'confusion_matrix' in results:
            results_with_cm[name] = {
                'confusion_matrix': results['confusion_matrix'],
                'accuracy': float(results.get('accuracy', 0)),
                'f1_macro': float(results.get('f1_macro', 0)),
            }

    if results_with_cm:
        save_path = os.path.join(FIGURES_DIR, 'confusion_matrices',
                                  'confusion_matrices_all.png')
        plot_confusion_matrices_comparison(results_with_cm, save_path=save_path)


def generate_training_curves():
    """Generate training curve comparison figures."""
    for name, model_dir in APPROACHES.items():
        history_path = os.path.join(RESULTS_DIR, 'models', model_dir,
                                     'training_history.npz')

        if not os.path.exists(history_path):
            continue

        data = np.load(history_path)
        history = {key: data[key].tolist() for key in data.files}

        save_path = os.path.join(FIGURES_DIR, 'training_curves',
                                  f'training_curves_{model_dir}.png')
        plot_training_curves(history, title=f'Training Curves: {name}',
                            save_path=save_path)


def generate_comparison_bar(all_results):
    """Generate approach comparison bar chart."""
    valid_results = {}
    for name, results in all_results.items():
        if results is not None:
            valid_results[name] = {
                'accuracy': float(results.get('accuracy', 0)),
                'f1_macro': float(results.get('f1_macro', 0)),
                'kappa': float(results.get('kappa', 0)),
            }

    if valid_results:
        save_path = os.path.join(FIGURES_DIR, 'statistical',
                                  'approach_comparison_bar.png')
        plot_approach_comparison_bar(valid_results, save_path=save_path)


def generate_feature_importance():
    """Generate feature importance figure for RF."""
    fi_path = os.path.join(RESULTS_DIR, 'models', 'rf_change',
                            'feature_importance.npz')

    if not os.path.exists(fi_path):
        return

    data = np.load(fi_path, allow_pickle=True)
    importances = list(zip(data['names'], data['values']))

    save_path = os.path.join(FIGURES_DIR, 'statistical',
                              'rf_feature_importance.png')
    plot_feature_importance(importances, title='RF Change Detection Feature Importance',
                           save_path=save_path)


def main():
    """Generate all publication comparison outputs."""
    print("=" * 60)
    print("GENERATE PUBLICATION COMPARISON")
    print("=" * 60)

    os.makedirs(TABLES_DIR, exist_ok=True)
    for subdir in ['confusion_matrices', 'training_curves', 'statistical']:
        os.makedirs(os.path.join(FIGURES_DIR, subdir), exist_ok=True)

    # Load all results
    print("\nLoading approach results...")
    all_results = {}
    for name, model_dir in APPROACHES.items():
        results = load_approach_results(name, model_dir)
        all_results[name] = results
        if results is not None:
            print(f"  {name}: Loaded (Acc={float(results.get('accuracy', 0)):.4f})")

    loaded = sum(1 for r in all_results.values() if r is not None)
    if loaded == 0:
        print("\nNo results found! Train models first:")
        print("  python scripts/train_all_approaches.py")
        return

    print(f"\n  Loaded {loaded}/{len(APPROACHES)} approaches")

    # Generate outputs
    print("\n--- Performance Tables ---")
    generate_performance_table(all_results)
    generate_per_class_table(all_results)

    print("\n--- Confusion Matrices ---")
    generate_confusion_matrices(all_results)

    print("\n--- Training Curves ---")
    generate_training_curves()

    print("\n--- Comparison Bar Chart ---")
    generate_comparison_bar(all_results)

    print("\n--- Feature Importance ---")
    generate_feature_importance()

    print("\n" + "=" * 60)
    print("PUBLICATION COMPARISON COMPLETE")
    print("=" * 60)
    print(f"  Tables: {TABLES_DIR}")
    print(f"  Figures: {FIGURES_DIR}")


if __name__ == '__main__':
    main()
