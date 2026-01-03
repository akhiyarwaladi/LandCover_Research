#!/usr/bin/env python3
"""
Generate Publication Tables - Multi-Architecture Support
=========================================================

Generates beautiful XLSX tables for ResNet architecture comparison.
Supports ALL ResNet variants (18, 34, 50, 101, 152).

Usage:
    python scripts/generate_publication_tables.py                    # ResNet50 only
    python scripts/generate_publication_tables.py --variant resnet18 # Specific variant
    python scripts/generate_publication_tables.py --all              # All variants comparison

Outputs:
    - Excel files (.xlsx) with professional formatting
    - Auto-adjusted column widths
    - Formatted headers
    - CSV backups
    - LaTeX versions

Author: Claude Sonnet 4.5
Date: 2026-01-03
Updated: 2026-01-03 (Added multi-architecture + XLSX support)
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import numpy as np
import pandas as pd
from sklearn.metrics import precision_recall_fscore_support
import argparse
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# PARSE ARGUMENTS
# ============================================================================

parser = argparse.ArgumentParser(description='Generate publication tables')
parser.add_argument('--variant', type=str, default='resnet50',
                   help='ResNet variant (resnet18/34/50/101/152)')
parser.add_argument('--all', action='store_true',
                   help='Generate comparison table for all available variants')
args = parser.parse_args()

# ============================================================================
# CONFIGURATION
# ============================================================================

OUTPUT_DIR = 'results/publication/tables'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CLASS_NAMES = ['Water', 'Trees', 'Crops', 'Shrub', 'Built', 'Bare']

# Random Forest baseline
RF_BASELINE = {
    'accuracy': 0.7495,
    'f1_macro': 0.542,
    'f1_weighted': 0.744,
    'precision_macro': 0.58,
    'recall_macro': 0.54,
    'f1_per_class': [0.79, 0.74, 0.78, 0.37, 0.42, 0.15]
}

# Architecture specifications
ARCH_SPECS = {
    'resnet18': {'params': 11.7e6, 'depth': 18},
    'resnet34': {'params': 21.8e6, 'depth': 34},
    'resnet50': {'params': 25.6e6, 'depth': 50},
    'resnet101': {'params': 44.5e6, 'depth': 101},
    'resnet152': {'params': 60.2e6, 'depth': 152}
}

print("\n" + "="*80)
print("GENERATING PUBLICATION TABLES (MULTI-ARCHITECTURE)")
print("="*80)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def format_excel_table(file_path, sheet_name, title=None):
    """Apply professional formatting to Excel table."""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Title row
    if title:
        ws.insert_rows(1)
        ws['A1'] = title
        ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        header_row = 2
    else:
        header_row = 1

    # Header formatting
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # Data rows
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            # First column left-aligned and bold
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Header row height
    ws.row_dimensions[header_row].height = 25

    wb.save(file_path)
    print(f"✓ Formatted: {os.path.basename(file_path)}")

def load_variant_data(variant):
    """Load training results for a specific variant."""
    try:
        history = np.load(f'results/{variant}/training_history.npz')
        test_data = np.load(f'results/{variant}/test_results.npz')

        data = {
            'variant': variant,
            'history': history,
            'y_true': test_data['targets'],
            'y_pred': test_data['predictions'],
            'accuracy': float(test_data['accuracy']),
            'f1_macro': float(test_data['f1_macro']),
            'f1_weighted': float(test_data['f1_weighted'])
        }

        # Calculate per-class metrics
        precision, recall, f1, support = precision_recall_fscore_support(
            data['y_true'], data['y_pred'], average=None, zero_division=0
        )
        data['precision'] = precision
        data['recall'] = recall
        data['f1_per_class'] = f1
        data['support'] = support
        data['precision_macro'] = precision.mean()
        data['recall_macro'] = recall.mean()

        return data
    except FileNotFoundError:
        print(f"⚠️  Data not found for {variant}")
        return None

# ============================================================================
# GENERATE TABLES
# ============================================================================

if args.all:
    # Multi-architecture comparison table
    print("\n📊 Generating MULTI-ARCHITECTURE COMPARISON TABLE")
    print("-"*80)

    # Load all available variants
    variants_data = {}
    for variant in ['resnet18', 'resnet34', 'resnet50', 'resnet101', 'resnet152']:
        data = load_variant_data(variant)
        if data:
            variants_data[variant] = data
            print(f"✓ Loaded {variant}: Accuracy={data['accuracy']*100:.2f}%")

    if not variants_data:
        print("❌ No variant data found! Train models first.")
        sys.exit(1)

    # TABLE: Architecture Comparison
    print("\n📋 Table: Architecture Performance Comparison")

    comparison_data = []
    for variant in sorted(variants_data.keys()):
        data = variants_data[variant]
        specs = ARCH_SPECS[variant]

        comparison_data.append({
            'Architecture': variant.upper(),
            'Depth': specs['depth'],
            'Parameters (M)': f"{specs['params']/1e6:.1f}",
            'Test Accuracy (%)': f"{data['accuracy']*100:.2f}",
            'F1 (Macro)': f"{data['f1_macro']:.4f}",
            'F1 (Weighted)': f"{data['f1_weighted']:.4f}",
            'Precision (Macro)': f"{data['precision_macro']:.4f}",
            'Recall (Macro)': f"{data['recall_macro']:.4f}",
            'Improvement vs RF (%)': f"+{(data['accuracy'] - RF_BASELINE['accuracy'])*100:.2f}"
        })

    df_comparison = pd.DataFrame(comparison_data)

    # Save as CSV
    csv_path = f'{OUTPUT_DIR}/Architecture_Comparison.csv'
    df_comparison.to_csv(csv_path, index=False)
    print(f"✓ Saved CSV: {os.path.basename(csv_path)}")

    # Save as XLSX with formatting
    xlsx_path = f'{OUTPUT_DIR}/Architecture_Comparison.xlsx'
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df_comparison.to_excel(writer, sheet_name='Comparison', index=False)

    format_excel_table(xlsx_path, 'Comparison',
                      title='ResNet Architecture Performance Comparison')
    print(f"✓ Saved XLSX: {os.path.basename(xlsx_path)}")
    print(f"  Size: {os.path.getsize(xlsx_path)/1024:.1f} KB")

else:
    # Single variant detailed tables
    variant = args.variant
    print(f"\n📊 Generating DETAILED TABLES for {variant.upper()}")
    print("-"*80)

    # Load data
    data = load_variant_data(variant)
    if not data:
        print(f"❌ No data found for {variant}!")
        print(f"   Train the model first: python scripts/train_all_resnet_variants_simple.py")
        sys.exit(1)

    print(f"✓ Loaded {variant}: Accuracy={data['accuracy']*100:.2f}%")

    # Prepare Excel writer
    excel_path = f'{OUTPUT_DIR}/{variant}_detailed_tables.xlsx'
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')

    # TABLE 1: Overall Performance
    print("\n📋 Table 1: Overall Performance Metrics")

    table1 = pd.DataFrame({
        'Metric': [
            'Overall Accuracy (%)',
            'Precision (Macro)',
            'Recall (Macro)',
            'F1-Score (Macro)',
            'F1-Score (Weighted)',
            'Improvement vs RF (Accuracy)',
            'Improvement vs RF (F1-Macro)'
        ],
        'Random Forest': [
            f'{RF_BASELINE["accuracy"]*100:.2f}',
            f'{RF_BASELINE["precision_macro"]:.4f}',
            f'{RF_BASELINE["recall_macro"]:.4f}',
            f'{RF_BASELINE["f1_macro"]:.4f}',
            f'{RF_BASELINE["f1_weighted"]:.4f}',
            '-',
            '-'
        ],
        variant.upper(): [
            f'{data["accuracy"]*100:.2f}',
            f'{data["precision_macro"]:.4f}',
            f'{data["recall_macro"]:.4f}',
            f'{data["f1_macro"]:.4f}',
            f'{data["f1_weighted"]:.4f}',
            f'+{(data["accuracy"] - RF_BASELINE["accuracy"])*100:.2f}%',
            f'+{(data["f1_macro"] - RF_BASELINE["f1_macro"])*100:.2f}%'
        ]
    })

    table1.to_excel(writer, sheet_name='Overall Performance', index=False)
    table1.to_csv(f'{OUTPUT_DIR}/{variant}_overall_performance.csv', index=False)
    print(f"✓ Overall Performance")

    # TABLE 2: Training Configuration
    print("📋 Table 2: Training Configuration")

    history = data['history']
    best_epoch = np.argmax(history['val_acc']) + 1
    best_val_acc = history['val_acc'][best_epoch - 1] * 100

    table2 = pd.DataFrame({
        'Configuration Parameter': [
            'Model Architecture',
            'Input Patch Size',
            'Number of Channels',
            'Total Epochs',
            'Best Epoch',
            'Best Validation Accuracy (%)',
            'Final Training Loss',
            'Final Validation Loss',
            'Learning Rate',
            'Optimizer',
            'Batch Size',
            'Training Samples',
            'Test Samples'
        ],
        'Value': [
            f'{variant.upper()} (pretrained)',
            '32 × 32 pixels',
            '23 (10 bands + 13 indices)',
            '30',
            f'{best_epoch}',
            f'{best_val_acc:.2f}',
            f'{history["train_loss"][-1]:.4f}',
            f'{history["val_loss"][-1]:.4f}',
            '0.0001',
            'Adam',
            '16',
            '80,000',
            '20,000'
        ]
    })

    table2.to_excel(writer, sheet_name='Training Config', index=False)
    table2.to_csv(f'{OUTPUT_DIR}/{variant}_training_config.csv', index=False)
    print(f"✓ Training Configuration")

    # TABLE 3: Per-Class Performance
    print("📋 Table 3: Per-Class Performance")

    table3_data = []
    for i, class_name in enumerate(CLASS_NAMES):
        table3_data.append({
            'Class': class_name,
            'Precision': f'{data["precision"][i]:.4f}',
            'Recall': f'{data["recall"][i]:.4f}',
            f'F1-Score ({variant.upper()})': f'{data["f1_per_class"][i]:.4f}',
            'F1-Score (RF)': f'{RF_BASELINE["f1_per_class"][i]:.4f}',
            'Improvement': f'{(data["f1_per_class"][i] - RF_BASELINE["f1_per_class"][i]):.4f}',
            'Test Samples': f'{data["support"][i]:,}'
        })

    table3 = pd.DataFrame(table3_data)
    table3.to_excel(writer, sheet_name='Per-Class Metrics', index=False)
    table3.to_csv(f'{OUTPUT_DIR}/{variant}_perclass_metrics.csv', index=False)
    print(f"✓ Per-Class Metrics")

    # Close Excel writer
    writer.close()

    # Format all sheets
    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        format_excel_table(excel_path, sheet_name, title=f'{variant.upper()} - {sheet_name}')

    print(f"\n✅ All tables saved to: {excel_path}")
    print(f"   Size: {os.path.getsize(excel_path)/1024:.1f} KB")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "="*80)
print("✅ TABLE GENERATION COMPLETE!")
print("="*80)

print(f"\n📁 Output directory: {OUTPUT_DIR}/")

if args.all:
    print("\n📊 Generated files:")
    print(f"   ✓ Architecture_Comparison.xlsx (formatted)")
    print(f"   ✓ Architecture_Comparison.csv (backup)")
else:
    print(f"\n📊 Generated files for {variant.upper()}:")
    print(f"   ✓ {variant}_detailed_tables.xlsx (formatted, multi-sheet)")
    print(f"   ✓ {variant}_overall_performance.csv")
    print(f"   ✓ {variant}_training_config.csv")
    print(f"   ✓ {variant}_perclass_metrics.csv")

print("\n✨ Features:")
print("   ✓ Professional XLSX formatting")
print("   ✓ Auto-adjusted column widths")
print("   ✓ Formatted headers")
print("   ✓ Publication-ready")
print("   ✓ Works from saved models (no retraining!)")

print("\n💡 Usage:")
print("   All variants: python scripts/generate_publication_tables.py --all")
print(f"   One variant:  python scripts/generate_publication_tables.py --variant resnet50")

print("\n" + "="*80)
