#!/usr/bin/env python3
"""
Generate Publication-Quality ResNet Comparison
===============================================

Creates comprehensive comparison visualizations and tables for all ResNet variants.

Outputs:
    - Performance comparison table (LaTeX + PNG)
    - Confusion matrices (all 4 models)
    - Per-class F1 comparison
    - Training curves comparison

Author: Claude Sonnet 4.5
Date: 2026-01-04
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix, classification_report, f1_score
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# EXCEL FORMATTING FUNCTIONS
# ============================================================================

def format_excel_table(file_path, header_row=1, theme='grayscale', merge_column=None):
    """
    Apply professional journal-ready formatting to Excel file.

    Args:
        file_path: Path to Excel file
        header_row: Row number containing headers (1-indexed)
        theme: 'grayscale', 'color', or 'minimal'
        merge_column: Column index (1-indexed) to merge repeated values
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Professional journal-ready themes
    themes = {
        'grayscale': {
            'header_fill': PatternFill(start_color="404040", end_color="404040", fill_type="solid"),
            'header_font': Font(name='Calibri', bold=True, color="FFFFFF", size=11),
            'alt_row_fill': PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid"),
            'data_font': Font(name='Calibri', size=10)
        },
        'professional': {
            'header_fill': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
            'header_font': Font(name='Calibri', bold=True, color="FFFFFF", size=11),
            'alt_row_fill': PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid"),
            'data_font': Font(name='Calibri', size=10)
        },
        'minimal': {
            'header_fill': PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
            'header_font': Font(name='Calibri', bold=True, color="000000", size=11),
            'alt_row_fill': None,
            'data_font': Font(name='Calibri', size=10)
        }
    }

    theme_style = themes.get(theme, themes['grayscale'])

    # Border styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_top_border = Border(
        top=Side(style='medium'), left=Side(style='thin'),
        right=Side(style='thin'), bottom=Side(style='thin')
    )

    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            # Border
            if cell.row == header_row:
                cell.border = thick_top_border
            else:
                cell.border = thin_border

            # Alignment
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Header row
            if cell.row == header_row:
                cell.fill = theme_style['header_fill']
                cell.font = theme_style['header_font']
            # Data rows
            else:
                cell.font = theme_style['data_font']
                # Alternating row colors
                if theme_style['alt_row_fill'] and cell.row % 2 == 0:
                    cell.fill = theme_style['alt_row_fill']

            # First column left-aligned and bold
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if cell.row != header_row:
                    cell.font = Font(name='Calibri', bold=True, size=10)

    # Merge repeated values in specified column
    if merge_column:
        merge_cells_by_value(ws, merge_column, header_row)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 3, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Header row height
    ws.row_dimensions[header_row].height = 25

    wb.save(file_path)
    print(f"✓ Formatted Excel ({theme}): {os.path.basename(file_path)}")


def merge_cells_by_value(ws, column_idx, header_row):
    """
    Merge cells with same consecutive values in specified column.

    Args:
        ws: Worksheet object
        column_idx: Column index (1-indexed)
        header_row: Header row number (1-indexed)
    """
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(column_idx)
    max_row = ws.max_row

    start_row = header_row + 1
    current_value = ws[f"{col_letter}{start_row}"].value
    merge_start = start_row

    for row in range(start_row + 1, max_row + 2):
        if row <= max_row:
            cell_value = ws[f"{col_letter}{row}"].value
        else:
            cell_value = None  # Trigger final merge

        if cell_value != current_value:
            # Merge cells if range > 1
            if row - 1 > merge_start:
                ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{row-1}")
                # Center the merged cell vertically
                ws[f"{col_letter}{merge_start}"].alignment = Alignment(
                    horizontal='left', vertical='center'
                )

            # Start new merge group
            merge_start = row
            current_value = cell_value

# ============================================================================
# CONFIGURATION
# ============================================================================

VARIANTS = ['resnet18', 'resnet34', 'resnet101', 'resnet152']
CLASS_NAMES = ['Water', 'Trees', 'Crops', 'Shrub', 'Built', 'Bare']

# New centralized paths
TABLES_DIR = 'results/tables/performance'
FIGURES_CONFUSION_DIR = 'results/figures/confusion_matrices'
FIGURES_TRAINING_DIR = 'results/figures/training_curves'

os.makedirs(TABLES_DIR, exist_ok=True)
os.makedirs(FIGURES_CONFUSION_DIR, exist_ok=True)
os.makedirs(FIGURES_TRAINING_DIR, exist_ok=True)

print("\n" + "="*80)
print("PUBLICATION-QUALITY RESNET COMPARISON")
print("="*80)

# ============================================================================
# LOAD ALL RESULTS
# ============================================================================

print("\n" + "-"*80)
print("Loading Results from All Variants")
print("-"*80)

all_results = {}
all_training = {}

for variant in VARIANTS:
    test_path = f'results/models/{variant}/test_results.npz'
    train_path = f'results/models/{variant}/training_history.npz'

    if os.path.exists(test_path):
        data = np.load(test_path)
        all_results[variant] = {
            'predictions': data['predictions'],
            'targets': data['targets'],
            'accuracy': float(data['accuracy']),
            'f1_macro': float(data['f1_macro']),
            'f1_weighted': float(data['f1_weighted'])
        }
        print(f"✓ {variant}: Acc={data['accuracy']:.4f}, F1={data['f1_macro']:.4f}")
    else:
        print(f"⚠️  {variant}: test_results.npz not found")

    if os.path.exists(train_path):
        all_training[variant] = np.load(train_path)

# ============================================================================
# 1. PERFORMANCE COMPARISON TABLE
# ============================================================================

print("\n" + "-"*80)
print("1/6: Generating Performance Comparison Table")
print("-"*80)

# Create DataFrame
table_data = []
for variant in VARIANTS:
    if variant in all_results:
        r = all_results[variant]
        table_data.append({
            'Model': variant.upper(),
            'Accuracy (%)': f"{r['accuracy']*100:.2f}",
            'F1-Macro': f"{r['f1_macro']:.4f}",
            'F1-Weighted': f"{r['f1_weighted']:.4f}"
        })

df = pd.DataFrame(table_data)

# Save as Excel with professional formatting
xlsx_path = os.path.join(TABLES_DIR, 'performance_table.xlsx')
df.to_excel(xlsx_path, index=False, sheet_name='Performance Comparison')
format_excel_table(xlsx_path, header_row=1, theme='professional')
print(f"✓ Saved Excel: {xlsx_path}")

# Save as LaTeX
latex_path = os.path.join(TABLES_DIR, 'performance_table.tex')
with open(latex_path, 'w') as f:
    f.write("\\begin{table}[h]\n")
    f.write("\\centering\n")
    f.write("\\caption{Performance Comparison of ResNet Architectures}\n")
    f.write("\\label{tab:resnet_comparison}\n")
    f.write("\\begin{tabular}{lccc}\n")
    f.write("\\hline\n")
    f.write("Model & Accuracy (\\%) & F1-Macro & F1-Weighted \\\\\n")
    f.write("\\hline\n")
    for _, row in df.iterrows():
        f.write(f"{row['Model']} & {row['Accuracy (%)']} & {row['F1-Macro']} & {row['F1-Weighted']} \\\\\n")
    f.write("\\hline\n")
    f.write("\\end{tabular}\n")
    f.write("\\end{table}\n")
print(f"✓ Saved LaTeX: {latex_path}")

# ============================================================================
# 2. CONFUSION MATRICES (All 4 models) - UNIQUE: Shows error patterns
# ============================================================================

print("\n" + "-"*80)
print("2/4: Generating Confusion Matrices (Error Pattern Analysis)")
print("-"*80)

fig, axes = plt.subplots(2, 2, figsize=(16, 14))
axes = axes.flatten()

for idx, variant in enumerate(VARIANTS):
    if variant not in all_results:
        continue

    r = all_results[variant]
    cm = confusion_matrix(r['targets'], r['predictions'])
    cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]

    ax = axes[idx]
    sns.heatmap(cm_normalized, annot=True, fmt='.2f', cmap='Blues',
                xticklabels=CLASS_NAMES, yticklabels=CLASS_NAMES,
                ax=ax, cbar_kws={'label': 'Proportion'})
    ax.set_title(f'{variant.upper()} - Acc: {r["accuracy"]:.2%}, F1: {r["f1_macro"]:.4f}',
                fontsize=12, fontweight='bold')
    ax.set_ylabel('True Label', fontsize=11)
    ax.set_xlabel('Predicted Label', fontsize=11)
    ax.tick_params(labelsize=10)

plt.tight_layout()
cm_path = os.path.join(FIGURES_CONFUSION_DIR, 'confusion_matrices_all.png')
plt.savefig(cm_path, dpi=300, bbox_inches='tight', facecolor='white')
plt.close()
print(f"✓ Saved: {cm_path}")

# ============================================================================
# 3. TRAINING CURVES - UNIQUE: Shows convergence over time
# ============================================================================

print("\n" + "-"*80)
print("3/4: Generating Training Curves (Convergence Analysis)")
print("-"*80)

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))

colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']

for idx, variant in enumerate(VARIANTS):
    if variant not in all_training:
        continue

    train_data = all_training[variant]
    epochs = range(1, len(train_data['train_loss']) + 1)

    # Loss curves
    ax1.plot(epochs, train_data['train_loss'],
            label=f'{variant.upper()} Train',
            color=colors[idx], linestyle='-', linewidth=2, alpha=0.8)
    ax1.plot(epochs, train_data['val_loss'],
            label=f'{variant.upper()} Val',
            color=colors[idx], linestyle='--', linewidth=2, alpha=0.6)

    # Accuracy curves
    ax2.plot(epochs, train_data['train_acc'],
            label=f'{variant.upper()} Train',
            color=colors[idx], linestyle='-', linewidth=2, alpha=0.8)
    ax2.plot(epochs, train_data['val_acc'],
            label=f'{variant.upper()} Val',
            color=colors[idx], linestyle='--', linewidth=2, alpha=0.6)

ax1.set_xlabel('Epoch', fontsize=12, fontweight='bold')
ax1.set_ylabel('Loss', fontsize=12, fontweight='bold')
ax1.set_title('Training and Validation Loss', fontsize=13, fontweight='bold')
ax1.legend(fontsize=9, loc='upper right')
ax1.grid(alpha=0.3)

ax2.set_xlabel('Epoch', fontsize=12, fontweight='bold')
ax2.set_ylabel('Accuracy (%)', fontsize=12, fontweight='bold')
ax2.set_title('Training and Validation Accuracy', fontsize=13, fontweight='bold')
ax2.legend(fontsize=9, loc='lower right')
ax2.grid(alpha=0.3)

plt.tight_layout()
curves_path = os.path.join(FIGURES_TRAINING_DIR, 'training_curves_comparison.png')
plt.savefig(curves_path, dpi=300, bbox_inches='tight', facecolor='white')
plt.close()
print(f"✓ Saved: {curves_path}")

# ============================================================================
# 4. PER-CLASS PERFORMANCE TABLES (Multiple Layouts)
# ============================================================================

print("\n" + "-"*80)
print("4/4: Generating Per-Class Performance Tables (Multiple Layouts)")
print("-"*80)

# Create detailed per-class table data (keep numeric for pivoting)
class_table_data_numeric = []
for variant in VARIANTS:
    if variant not in all_results:
        continue

    r = all_results[variant]
    report = classification_report(r['targets'], r['predictions'],
                                  target_names=CLASS_NAMES, output_dict=True)

    for cls in CLASS_NAMES:
        class_table_data_numeric.append({
            'Model': variant.upper(),
            'Class': cls,
            'Precision': report[cls]['precision'],
            'Recall': report[cls]['recall'],
            'F1-Score': report[cls]['f1-score'],
            'Support': int(report[cls]['support'])
        })

df_class_numeric = pd.DataFrame(class_table_data_numeric)

# LAYOUT 1: Long format with merged cells (formatted for display)
df_class_formatted = df_class_numeric.copy()
df_class_formatted['Precision'] = df_class_formatted['Precision'].apply(lambda x: f"{x:.3f}")
df_class_formatted['Recall'] = df_class_formatted['Recall'].apply(lambda x: f"{x:.3f}")
df_class_formatted['F1-Score'] = df_class_formatted['F1-Score'].apply(lambda x: f"{x:.3f}")

class_long_path = os.path.join(TABLES_DIR, 'per_class_detailed_long.xlsx')
df_class_formatted.to_excel(class_long_path, index=False, sheet_name='Per-Class Performance')
format_excel_table(class_long_path, header_row=1, theme='professional', merge_column=1)
print(f"✓ Long format (merged): {os.path.basename(class_long_path)}")

# LAYOUT 2: Transposed format for each metric (compact for journals)
# Precision table (Classes as rows, Models as columns)
df_precision = df_class_numeric.pivot_table(index='Class', columns='Model', values='Precision', aggfunc='first')
precision_path = os.path.join(TABLES_DIR, 'per_class_precision_transposed.xlsx')
df_precision.to_excel(precision_path, sheet_name='Precision by Class')
format_excel_table(precision_path, header_row=1, theme='professional')
print(f"✓ Precision transposed: {os.path.basename(precision_path)}")

# Recall table
df_recall = df_class_numeric.pivot_table(index='Class', columns='Model', values='Recall', aggfunc='first')
recall_path = os.path.join(TABLES_DIR, 'per_class_recall_transposed.xlsx')
df_recall.to_excel(recall_path, sheet_name='Recall by Class')
format_excel_table(recall_path, header_row=1, theme='professional')
print(f"✓ Recall transposed: {os.path.basename(recall_path)}")

# F1-Score table
df_f1 = df_class_numeric.pivot_table(index='Class', columns='Model', values='F1-Score', aggfunc='first')
f1_path = os.path.join(TABLES_DIR, 'per_class_f1_transposed.xlsx')
df_f1.to_excel(f1_path, sheet_name='F1-Score by Class')
format_excel_table(f1_path, header_row=1, theme='professional')
print(f"✓ F1-Score transposed: {os.path.basename(f1_path)}")

# LAYOUT 3: Comprehensive transposed (all metrics, very compact)
# Create multi-index for Class + Metric
metrics_compact = []
for metric in ['Precision', 'Recall', 'F1-Score']:
    df_metric = df_class_numeric.pivot_table(index='Class', columns='Model', values=metric, aggfunc='first')
    df_metric['Metric'] = metric
    metrics_compact.append(df_metric)

df_compact = pd.concat(metrics_compact)
df_compact = df_compact.reset_index().set_index(['Class', 'Metric'])
compact_path = os.path.join(TABLES_DIR, 'per_class_all_metrics_compact.xlsx')
df_compact.to_excel(compact_path, sheet_name='All Metrics Compact')
format_excel_table(compact_path, header_row=1, theme='professional')
print(f"✓ All metrics compact: {os.path.basename(compact_path)}")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "="*80)
print("PUBLICATION COMPARISON COMPLETE!")
print("="*80)

print(f"\n📁 Files saved to centralized structure:")
print(f"  📊 Tables: {TABLES_DIR}/")
print(f"  📈 Figures: results/figures/ (confusion_matrices/, training_curves/)")

print("\n📊 TABLES - Professional journal-ready formatting:")
print("\nOverall Performance:")
print("  1. performance_table.xlsx - Overall metrics (Accuracy, F1-Macro, F1-Weighted)")
print("  2. performance_table.tex - LaTeX format (for journal submission)")

print("\nPer-Class Performance (Multiple Layouts):")
print("  3. per_class_detailed_long.xlsx - Long format with merged cells (24 rows)")
print("  4. per_class_precision_transposed.xlsx - Compact (6 classes × 4 models)")
print("  5. per_class_recall_transposed.xlsx - Compact (6 classes × 4 models)")
print("  6. per_class_f1_transposed.xlsx - Compact (6 classes × 4 models)")
print("  7. per_class_all_metrics_compact.xlsx - Ultra-compact (all metrics, 18 rows)")

print("\n📈 FIGURES (for patterns & relationships):")
print("  8. confusion_matrices_all.png - ERROR PATTERNS (which classes confused)")
print("  9. training_curves_comparison.png - CONVERGENCE (learning over time)")

print("\n✨ Professional Theme:")
print("  • Calibri font (journal standard)")
print("  • Dark gray headers (#2C3E50)")
print("  • Alternating row colors for readability")
print("  • Merged cells for repeated model names")
print("  • Transposed layouts fit journal page widths")

print("\n✨ No redundancy - Tables show exact values, Figures show patterns!")
print("✨ Multiple layouts: Choose based on journal requirements!")

print("\n" + "="*80)
