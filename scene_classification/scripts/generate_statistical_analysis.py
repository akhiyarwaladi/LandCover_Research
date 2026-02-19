"""
Generate Statistical Analysis (McNemar's Tests, Kappa, Efficiency)

Usage:
    python scripts/generate_statistical_analysis.py
"""

import os
import sys
import json
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import RESULTS_DIR, MODELS
from modules.evaluator import mcnemar_test


def format_excel(writer, sheet_name, df):
    """Apply professional formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df.to_excel(writer, sheet_name=sheet_name, index=True)
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                              fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)


def generate_mcnemar_tables():
    """Generate McNemar's test tables for each dataset."""
    from modules.visualizer import plot_mcnemar_matrix

    table_dir = os.path.join(RESULTS_DIR, 'tables', 'statistical')
    fig_dir = os.path.join(RESULTS_DIR, 'figures', 'statistical')
    os.makedirs(table_dir, exist_ok=True)
    os.makedirs(fig_dir, exist_ok=True)

    models_dir = os.path.join(RESULTS_DIR, 'models')
    if not os.path.isdir(models_dir):
        print("No model results found.")
        return

    for ds_name in os.listdir(models_dir):
        ds_path = os.path.join(models_dir, ds_name)
        if not os.path.isdir(ds_path):
            continue

        # Load predictions for all models
        model_preds = {}
        y_true = None
        for m_name in os.listdir(ds_path):
            test_path = os.path.join(ds_path, m_name, 'test_results.npz')
            if os.path.exists(test_path):
                npz = np.load(test_path)
                model_preds[m_name] = npz['y_pred']
                if y_true is None:
                    y_true = npz['y_true']

        if len(model_preds) < 2 or y_true is None:
            continue

        # Pairwise McNemar
        model_names = sorted(model_preds.keys())
        rows = []
        comparisons = []
        for i in range(len(model_names)):
            for j in range(i + 1, len(model_names)):
                a, b = model_names[i], model_names[j]
                chi2_stat, p_val = mcnemar_test(
                    y_true, model_preds[a], model_preds[b])

                sig = 'ns'
                if p_val < 0.001:
                    sig = '***'
                elif p_val < 0.01:
                    sig = '**'
                elif p_val < 0.05:
                    sig = '*'

                rows.append({
                    'Model A': a,
                    'Model B': b,
                    'Chi-squared': round(chi2_stat, 4),
                    'p-value': round(p_val, 6),
                    'Significance': sig,
                })
                comparisons.append({
                    'model_a': a, 'model_b': b,
                    'chi2': chi2_stat, 'p_value': p_val,
                })

        df = pd.DataFrame(rows)
        path = os.path.join(table_dir, f'mcnemar_{ds_name}.xlsx')
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            format_excel(writer, 'McNemar', df.set_index('Model A'))
        print(f"  Table: {path}")

        # P-value heatmap
        plot_mcnemar_matrix(
            comparisons, model_names,
            save_path=os.path.join(fig_dir, f'mcnemar_{ds_name}.png'))


def generate_kappa_table():
    """Generate Kappa coefficient table."""
    table_dir = os.path.join(RESULTS_DIR, 'tables', 'statistical')
    os.makedirs(table_dir, exist_ok=True)

    models_dir = os.path.join(RESULTS_DIR, 'models')
    if not os.path.isdir(models_dir):
        return

    rows = []
    for ds_name in os.listdir(models_dir):
        ds_path = os.path.join(models_dir, ds_name)
        if not os.path.isdir(ds_path):
            continue
        for m_name in os.listdir(ds_path):
            mp = os.path.join(ds_path, m_name, 'evaluation_metrics.json')
            if os.path.exists(mp):
                with open(mp) as f:
                    metrics = json.load(f)
                kappa = metrics.get('kappa', 0)
                # Interpretation (Landis & Koch, 1977)
                if kappa >= 0.81:
                    interp = 'Almost Perfect'
                elif kappa >= 0.61:
                    interp = 'Substantial'
                elif kappa >= 0.41:
                    interp = 'Moderate'
                elif kappa >= 0.21:
                    interp = 'Fair'
                else:
                    interp = 'Slight/Poor'

                rows.append({
                    'Dataset': ds_name,
                    'Model': m_name,
                    'Kappa': round(kappa, 4),
                    'Interpretation': interp,
                    'Accuracy': round(metrics.get('accuracy', 0), 4),
                })

    if rows:
        df = pd.DataFrame(rows).set_index(['Dataset', 'Model'])
        path = os.path.join(table_dir, 'kappa_analysis.xlsx')
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            format_excel(writer, 'Kappa', df)
        print(f"  Table: {path}")


def generate_efficiency_table():
    """Generate computational efficiency table."""
    table_dir = os.path.join(RESULTS_DIR, 'tables', 'statistical')
    os.makedirs(table_dir, exist_ok=True)

    summary_path = os.path.join(RESULTS_DIR, 'all_experiments_summary.json')
    if not os.path.exists(summary_path):
        return

    with open(summary_path) as f:
        data = json.load(f)

    rows = []
    for ds_name, model_results in data.get('results', {}).items():
        for m_name, res in model_results.items():
            if 'error' in res:
                continue
            family = MODELS.get(m_name, {}).get('family', 'unknown')
            rows.append({
                'Dataset': ds_name,
                'Model': m_name,
                'Family': family,
                'Params (M)': round(res.get('params_m', 0), 1),
                'Training Time (s)': round(res.get('training_time', 0), 1),
                'Best Epoch': res.get('best_epoch', 0),
                'Accuracy': round(res.get('accuracy', 0), 4),
            })

    if rows:
        df = pd.DataFrame(rows).set_index(['Dataset', 'Model'])
        path = os.path.join(table_dir, 'computational_efficiency.xlsx')
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            format_excel(writer, 'Efficiency', df)
        print(f"  Table: {path}")


def main():
    print("=" * 60)
    print("STATISTICAL ANALYSIS")
    print("=" * 60)

    print("\nMcNemar's Tests...")
    generate_mcnemar_tables()

    print("\nKappa Analysis...")
    generate_kappa_table()

    print("\nComputational Efficiency...")
    generate_efficiency_table()

    print("\nDone!")


if __name__ == '__main__':
    main()
