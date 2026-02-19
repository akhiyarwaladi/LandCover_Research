"""
Generate Publication-Ready Tables and Figures

Produces Excel tables and PNG figures from experiment results.

Usage:
    python scripts/generate_publication_outputs.py
"""

import os
import sys
import json
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import RESULTS_DIR, MODELS, DATASETS
from modules.visualizer import (
    plot_confusion_matrices_grid, plot_training_curves,
    plot_accuracy_comparison, plot_per_class_f1, plot_model_efficiency
)


def format_excel(writer, sheet_name, df):
    """Apply professional formatting to an Excel sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df.to_excel(writer, sheet_name=sheet_name, index=True)
    ws = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                              fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)


def load_results():
    """Load all experiment results."""
    summary_path = os.path.join(RESULTS_DIR, 'all_experiments_summary.json')
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            return json.load(f)

    # Try to reconstruct from individual results
    results = {}
    models_dir = os.path.join(RESULTS_DIR, 'models')
    if not os.path.isdir(models_dir):
        return None

    for ds_name in os.listdir(models_dir):
        ds_path = os.path.join(models_dir, ds_name)
        if not os.path.isdir(ds_path):
            continue
        results[ds_name] = {}
        for model_name in os.listdir(ds_path):
            metrics_path = os.path.join(ds_path, model_name,
                                        'evaluation_metrics.json')
            if os.path.exists(metrics_path):
                with open(metrics_path) as f:
                    metrics = json.load(f)
                results[ds_name][model_name] = metrics

    return {'results': results}


def generate_performance_tables(data):
    """Generate performance comparison Excel tables."""
    results = data['results']
    table_dir = os.path.join(RESULTS_DIR, 'tables', 'performance')
    os.makedirs(table_dir, exist_ok=True)

    for ds_name, model_results in results.items():
        if not model_results:
            continue

        # Overall performance table
        rows = []
        for m_name, m_res in model_results.items():
            if 'error' in m_res:
                continue
            rows.append({
                'Model': m_name,
                'Accuracy': m_res.get('accuracy', 0),
                'F1-Macro': m_res.get('f1_macro', 0),
                'F1-Weighted': m_res.get('f1_weighted', 0),
                'Kappa': m_res.get('kappa', 0),
                'Params (M)': m_res.get('params_m', 0),
                'Time (s)': m_res.get('training_time', 0),
            })

        if not rows:
            continue

        df = pd.DataFrame(rows).set_index('Model').sort_values(
            'Accuracy', ascending=False)

        path = os.path.join(table_dir, f'performance_{ds_name}.xlsx')
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            format_excel(writer, 'Performance', df)
        print(f"  Table: {path}")

    # Cross-dataset summary
    if len(results) > 1:
        rows = []
        for m_name in MODELS:
            row = {'Model': m_name}
            for ds_name in results:
                if m_name in results[ds_name] and 'error' not in results[ds_name][m_name]:
                    row[f'{ds_name}_acc'] = results[ds_name][m_name]['accuracy']
                    row[f'{ds_name}_f1'] = results[ds_name][m_name]['f1_macro']
            if len(row) > 1:
                rows.append(row)

        if rows:
            df = pd.DataFrame(rows).set_index('Model')
            path = os.path.join(table_dir, 'cross_dataset_comparison.xlsx')
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                format_excel(writer, 'Cross-Dataset', df)
            print(f"  Table: {path}")


def generate_per_class_tables(data):
    """Generate per-class performance tables."""
    results = data['results']
    table_dir = os.path.join(RESULTS_DIR, 'tables', 'performance')
    os.makedirs(table_dir, exist_ok=True)

    for ds_name, model_results in results.items():
        # Load detailed per-class from individual evaluations
        rows = []
        for m_name in model_results:
            metrics_path = os.path.join(RESULTS_DIR, 'models', ds_name,
                                        m_name, 'evaluation_metrics.json')
            if not os.path.exists(metrics_path):
                continue
            with open(metrics_path) as f:
                metrics = json.load(f)

            class_names = metrics.get('class_names', [])
            per_class = metrics.get('per_class', {})
            for i, cls in enumerate(class_names):
                rows.append({
                    'Model': m_name,
                    'Class': cls,
                    'Precision': per_class['precision'][i],
                    'Recall': per_class['recall'][i],
                    'F1': per_class['f1'][i],
                    'Support': per_class['support'][i],
                })

        if rows:
            df = pd.DataFrame(rows)
            path = os.path.join(table_dir, f'per_class_{ds_name}.xlsx')
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                format_excel(writer, 'Per-Class', df.set_index(['Model', 'Class']))
            print(f"  Table: {path}")


def generate_figures(data):
    """Generate all publication figures."""
    results = data['results']
    fig_dir = os.path.join(RESULTS_DIR, 'figures')

    for ds_name, model_results in results.items():
        if not model_results:
            continue

        # Load evaluation data for confusion matrices
        eval_data = {}
        histories = {}
        for m_name in model_results:
            test_path = os.path.join(RESULTS_DIR, 'models', ds_name,
                                     m_name, 'test_results.npz')
            hist_path = os.path.join(RESULTS_DIR, 'models', ds_name,
                                     m_name, 'training_history.npz')

            if os.path.exists(test_path):
                npz = np.load(test_path, allow_pickle=True)
                metrics_path = os.path.join(RESULTS_DIR, 'models', ds_name,
                                            m_name, 'evaluation_metrics.json')
                with open(metrics_path) as f:
                    metrics = json.load(f)
                eval_data[m_name] = {
                    'y_true': npz['y_true'],
                    'y_pred': npz['y_pred'],
                    'accuracy': metrics['accuracy'],
                    'per_class': metrics['per_class'],
                }

            if os.path.exists(hist_path):
                npz = np.load(hist_path)
                histories[m_name] = {
                    'train_loss': npz['train_loss'],
                    'train_acc': npz['train_acc'],
                    'test_loss': npz['test_loss'],
                    'test_acc': npz['test_acc'],
                }

        # Get class names
        class_names = None
        for m_name in model_results:
            mp = os.path.join(RESULTS_DIR, 'models', ds_name,
                              m_name, 'evaluation_metrics.json')
            if os.path.exists(mp):
                with open(mp) as f:
                    class_names = json.load(f).get('class_names')
                break

        if not class_names:
            continue

        # Confusion matrices
        if eval_data:
            plot_confusion_matrices_grid(
                eval_data, class_names, ds_name,
                save_path=os.path.join(fig_dir, 'confusion_matrices',
                                       f'cm_{ds_name}.png'))

        # Training curves
        if histories:
            plot_training_curves(
                histories,
                save_path=os.path.join(fig_dir, 'training_curves',
                                       f'curves_{ds_name}.png'))

        # Per-class F1
        if eval_data:
            plot_per_class_f1(
                eval_data, class_names, ds_name,
                save_path=os.path.join(fig_dir, 'statistical',
                                       f'per_class_f1_{ds_name}.png'))

    # Cross-dataset accuracy comparison
    acc_summary = {}
    for ds_name in results:
        acc_summary[ds_name] = {}
        for m_name in results[ds_name]:
            if 'error' not in results[ds_name][m_name]:
                acc_summary[ds_name][m_name] = results[ds_name][m_name].get(
                    'accuracy', 0)
    if acc_summary:
        plot_accuracy_comparison(
            acc_summary,
            save_path=os.path.join(fig_dir, 'statistical',
                                   'accuracy_comparison.png'))

    # Model efficiency plot (for first dataset with results)
    for ds_name in results:
        accs = {m: r.get('accuracy', 0) for m, r in results[ds_name].items()
                if 'error' not in r}
        if accs:
            plot_model_efficiency(
                MODELS, accs,
                save_path=os.path.join(fig_dir, 'statistical',
                                       f'efficiency_{ds_name}.png'))
            break


def main():
    print("=" * 60)
    print("GENERATE PUBLICATION OUTPUTS")
    print("=" * 60)

    data = load_results()
    if not data or not data.get('results'):
        print("No results found. Run train_all_experiments.py first.")
        return

    print("\nGenerating tables...")
    generate_performance_tables(data)
    generate_per_class_tables(data)

    print("\nGenerating figures...")
    generate_figures(data)

    print("\nDone!")


if __name__ == '__main__':
    main()
